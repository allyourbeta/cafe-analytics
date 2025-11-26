#!/usr/bin/env python3
"""
Extract products from database that are NOT in the provided Excel file.
Outputs a new Excel file with the missing products.

Usage:
    python extract_missing_products.py <existing_excel_file> <output_excel_file>

Example:
    python extract_missing_products.py products_168.xlsx missing_products.xlsx
"""

import sqlite3
import sys
import openpyxl
from pathlib import Path


def load_existing_product_ids(excel_path):
    """Load product IDs from existing Excel file."""
    print(f"\n📄 Reading existing Excel file: {excel_path}")
    
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    
    existing_ids = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:  # product_id column
            existing_ids.add(int(row[0]))
    
    print(f"  ✅ Found {len(existing_ids)} existing products")
    return existing_ids


def load_all_database_products(db_path):
    """Load all products from the database."""
    print(f"\n💾 Reading database: {db_path}")
    
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    
    cursor.execute("""
        SELECT item_id, item_name, category, current_price, current_cost, is_resold
        FROM items
        ORDER BY item_id
    """)
    
    products = cursor.fetchall()
    conn.close()
    
    print(f"  ✅ Found {len(products)} total products in database")
    return products


def write_missing_products_to_excel(missing_products, output_path):
    """Write missing products to a new Excel file."""
    print(f"\n📝 Writing missing products to: {output_path}")
    
    # Create new workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Missing Products"
    
    # Write headers
    headers = ['product_id', 'product_name', 'category', 'current_price', 'current_cost', 'is_resold']
    ws.append(headers)
    
    # Write data rows
    for product in missing_products:
        item_id, item_name, category, current_price, current_cost, is_resold = product
        ws.append([item_id, item_name, category, current_price, current_cost, is_resold])
    
    # Save workbook
    wb.save(output_path)
    print(f"  ✅ Wrote {len(missing_products)} products")


def main(existing_excel_path, output_excel_path):
    """Main extraction logic."""
    
    # Validate inputs
    if not Path(existing_excel_path).exists():
        print(f"❌ Error: File not found: {existing_excel_path}")
        sys.exit(1)
    
    db_path = Path(__file__).parent.parent / "database" / "cafe_reports.db"
    if not db_path.exists():
        print(f"❌ Error: Database not found: {db_path}")
        sys.exit(1)
    
    print("🚀 Extract Missing Products from Database")
    print(f"   Existing Excel: {existing_excel_path}")
    print(f"   Output Excel: {output_excel_path}")
    print(f"   Database: {db_path}")
    
    # Load data
    existing_ids = load_existing_product_ids(existing_excel_path)
    all_products = load_all_database_products(str(db_path))
    
    # Filter to missing products
    print(f"\n🔍 Identifying missing products...")
    missing_products = [
        product for product in all_products 
        if product[0] not in existing_ids  # product[0] is item_id
    ]
    
    print(f"  ✅ Found {len(missing_products)} products NOT in Excel file")
    print(f"  ✅ Found {len(all_products) - len(missing_products)} products already in Excel file")
    
    # Write output
    write_missing_products_to_excel(missing_products, output_excel_path)
    
    print(f"\n✅ Complete!")
    print(f"   Missing products saved to: {output_excel_path}")
    print(f"   You can now review and append these to your main spreadsheet.")


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python extract_missing_products.py <existing_excel_file> <output_excel_file>")
        print("\nExample:")
        print("  python extract_missing_products.py products_168.xlsx missing_products.xlsx")
        sys.exit(1)
    
    existing_excel = sys.argv[1]
    output_excel = sys.argv[2]
    
    main(existing_excel, output_excel)
